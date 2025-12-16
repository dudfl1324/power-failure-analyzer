import json
import os
import queue
import random
import sys
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd
import pyvisa
import tkinter as tk
import tkinter.font as tkfont
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from openpyxl import Workbook, load_workbook
from tkinter import filedialog, messagebox, ttk


APP_TITLE_LOGIN = "AutoPowerTester Login"
APP_TITLE_RESULTS = "Test Result"

MAIN_WINDOW_GEOMETRY = "1500x820"
CONFIG_DIALOG_GEOMETRY = "980x620"
PLOT_WINDOW_GEOMETRY = "700x500"

SAMPLE_COUNT = 30
NO_CURRENT_SECONDS = 15
CURRENT_PRESENT_THRESHOLD_A = 0.005

RESULTS_HEADER_FONT_SIZE = 12
RESULTS_BODY_FONT_SIZE = 8

DEFAULT_BOOT_VOLTAGE = 5.5

ADMIN_USERNAME = "admin"
DEV_USERNAME = "dev"
ADMIN_PASSWORD = "@Eng2019"
DEV_PASSWORD = "@Eng2019"

SUMMARY_HEADERS = ["IMEI", "Model", "Avg. Current", "P/F", "Worker ID", "Power Supply", "Date"]
DETAIL_HEADERS = ["IMEI", "Model", "Worker ID", "Power Supply", "Date", "SampleIndex", "Current(A)"]

DEFAULT_PSU_NAME = "PSU1"
DEFAULT_PSU_ADDRESS = "GPIB0::1::INSTR"


def get_script_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)

    try:
        return os.path.dirname(os.path.abspath(__file__))
    except NameError:
        return os.getcwd()


SCRIPT_DIR = get_script_dir()
CONFIG_FILE = os.path.join(SCRIPT_DIR, "AutoPowerTester_Configuration.json")

DOCUMENTS_LOG_DIR = os.path.join(os.path.expanduser("~"), r"Documents\AutoPowerTester Log")
os.makedirs(DOCUMENTS_LOG_DIR, exist_ok=True)


def _show_startup_error(title: str, message: str) -> None:
    temp_root = tk.Tk()
    temp_root.withdraw()
    messagebox.showerror(title, message)
    temp_root.destroy()


def _normalize_supply_entry(entry: Any, idx: int) -> Dict[str, str]:
    default_name = f"PSU{idx + 1}"
    if not isinstance(entry, dict):
        return {"name": default_name, "address": ""}

    name = str(entry.get("name", default_name)).strip() or default_name
    address = str(entry.get("address", "")).strip()
    return {"name": name, "address": address}


def _normalize_power_supplies(ps_list: Any) -> List[Dict[str, str]]:
    normalized: List[Dict[str, str]] = []
    if isinstance(ps_list, list):
        for i, entry in enumerate(ps_list[:4]):
            normalized.append(_normalize_supply_entry(entry, i))

    if not normalized:
        normalized = [{"name": DEFAULT_PSU_NAME, "address": DEFAULT_PSU_ADDRESS}]

    return normalized[:4]


def _warn_if_missing_psu_addresses(ps_list: List[Dict[str, str]]) -> None:
    missing = [ps.get("name", "PSU") for ps in ps_list if not str(ps.get("address", "")).strip()]
    if not missing:
        return

    messagebox.showwarning(
        "Power Supply Address Missing",
        "One or more power supply addresses are missing/empty:\n"
        f"- {', '.join(missing)}\n\n"
        "The program will still run, but real measurements may fail to connect.\n"
        "Please update AutoPowerTester_Configuration.json -> POWER_SUPPLIES.",
    )


def load_config() -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, str], List[Dict[str, str]]]:
    if not os.path.exists(CONFIG_FILE):
        _show_startup_error(
            "Missing Configuration",
            'Put "AutoPowerTester_Configuration.json" file in the same folder as AutoPowerTester.exe',
        )
        raise SystemExit(1)

    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    try:
        model_voltage_map = data["MODEL_VOLTAGE_MAP"]
        model_criteria = data["MODEL_CRITERIA"]
    except KeyError as e:
        _show_startup_error(
            "Invalid Configuration",
            f'Missing key in "AutoPowerTester_Configuration.json": {e}\n\n'
            'Make sure it contains "MODEL_VOLTAGE_MAP" and "MODEL_CRITERIA".',
        )
        raise SystemExit(1)

    workers = data.get("WORKERS", {})

    if "POWER_SUPPLIES" in data:
        ps_list_raw = data.get("POWER_SUPPLIES", [])
        ps_list = _normalize_power_supplies(ps_list_raw)
    else:
        addr = str(data.get("POWER_SUPPLY_ADDRESS", DEFAULT_PSU_ADDRESS)).strip() or DEFAULT_PSU_ADDRESS
        ps_list = _normalize_power_supplies([{"name": DEFAULT_PSU_NAME, "address": addr}])

    if not ps_list:
        ps_list = _normalize_power_supplies([])

    return model_voltage_map, model_criteria, workers, ps_list


def save_config(
    model_voltage_map: Dict[str, Any],
    model_criteria: Dict[str, Any],
    workers: Dict[str, str],
    power_supplies: List[Dict[str, str]],
) -> None:
    data = {
        "MODEL_VOLTAGE_MAP": model_voltage_map,
        "MODEL_CRITERIA": model_criteria,
        "WORKERS": workers,
        "POWER_SUPPLIES": power_supplies,
    }
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


MODEL_VOLTAGE_MAP, MODEL_CRITERIA, WORKERS, POWER_SUPPLIES = load_config()

try:
    _tmp = tk.Tk()
    _tmp.withdraw()
    _warn_if_missing_psu_addresses(POWER_SUPPLIES)
    _tmp.destroy()
except Exception:
    pass

RUN_SERIES_BY_ROW: Dict[str, Dict[str, Any]] = {}
LOG_WRITE_LOCK = threading.Lock()

PANEL_STATE: List[Dict[str, Optional[str]]] = [
    {"last_model": None, "last_imei": None},
    {"last_model": None, "last_imei": None},
    {"last_model": None, "last_imei": None},
    {"last_model": None, "last_imei": None},
]


@dataclass
class PanelJob:
    panel_index: int
    thread: threading.Thread
    stop_event: threading.Event
    series_token: str


ACTIVE_JOBS: List[Optional[PanelJob]] = [None, None, None, None]

SERIES_BY_TOKEN: Dict[str, Dict[str, List[float]]] = {}
SERIES_LOCK = threading.Lock()

log_saved_path: Optional[str] = None


def get_daily_log_saved_path() -> str:
    global log_saved_path
    today = datetime.now().strftime("%m%d%y")
    log_saved_path = os.path.join(DOCUMENTS_LOG_DIR, f"AutoPowerTester_Log_{today}.xlsx")
    return log_saved_path


def _ensure_sheet_headers(ws, headers: Sequence[str]) -> None:
    if ws.max_row < 1:
        for j, h in enumerate(headers, start=1):
            ws.cell(row=1, column=j, value=h)
        return

    first_row = [ws.cell(row=1, column=j).value for j in range(1, len(headers) + 1)]
    if all(v is None for v in first_row):
        for j, h in enumerate(headers, start=1):
            ws.cell(row=1, column=j, value=h)


def append_run_to_daily_log(summary_row: Sequence[Any], detail_rows: Sequence[Sequence[Any]]) -> None:
    path = get_daily_log_saved_path()

    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        _ensure_sheet_headers(ws, SUMMARY_HEADERS)
        ws2 = wb.create_sheet("Detailed")
        _ensure_sheet_headers(ws2, DETAIL_HEADERS)
        wb.save(path)

    wb = load_workbook(path)

    if "Summary" not in wb.sheetnames:
        ws = wb.create_sheet("Summary", 0)
    else:
        ws = wb["Summary"]
    _ensure_sheet_headers(ws, SUMMARY_HEADERS)

    if "Detailed" not in wb.sheetnames:
        ws2 = wb.create_sheet("Detailed")
    else:
        ws2 = wb["Detailed"]
    _ensure_sheet_headers(ws2, DETAIL_HEADERS)

    ws.append(list(summary_row))
    for r in detail_rows:
        ws2.append(list(r))

    wb.save(path)


def check_daily_log_file_not_open() -> str:
    path = get_daily_log_saved_path()
    if os.path.exists(path):
        with open(path, "rb+"):
            pass
    return path


def show_log_file_open_error(parent: tk.Misc, path: str) -> None:
    win = tk.Toplevel(parent)
    win.title("Log File Open / Archivo Abierto")
    try:
        win.attributes("-topmost", True)
    except Exception:
        pass

    win.resizable(False, False)

    dialog_w, dialog_h = 900, 320
    try:
        win.update_idletasks()
        sx = win.winfo_screenwidth()
        sy = win.winfo_screenheight()
        x = max(int((sx - dialog_w) / 2), 0)
        y = max(int((sy - dialog_h) / 2), 0)
        win.geometry(f"{dialog_w}x{dialog_h}+{x}+{y}")
    except Exception:
        win.geometry(f"{dialog_w}x{dialog_h}")

    container = tk.Frame(win, padx=18, pady=16)
    container.pack(fill="both", expand=True)

    tk.Label(container, text="LOG FILE IS OPEN", font=("TkDefaultFont", 18, "bold"), fg="red").pack(pady=(0, 10))
    tk.Label(
        container,
        text="The log file is open. Close the log file and test the device again!",
        font=("TkDefaultFont", 16),
        wraplength=840,
        justify="center",
    ).pack(pady=(0, 8))
    tk.Label(
        container,
        text="El archivo de registro está abierto. Cierre el archivo de registro y vuelva a probar el dispositivo.",
        font=("TkDefaultFont", 16),
        wraplength=840,
        justify="center",
    ).pack(pady=(0, 12))
    tk.Label(container, text=f"File: {path}", font=("TkDefaultFont", 11), fg="gray").pack(pady=(0, 16))

    def close() -> None:
        win.destroy()

    tk.Button(container, text="OK", font=("TkDefaultFont", 16, "bold"), width=10, command=close).pack()
    win.protocol("WM_DELETE_WINDOW", close)
    win.grab_set()
    try:
        win.focus_force()
    except Exception:
        pass

    parent.wait_window(win)


USE_PSEUDO_CURRENT = False
PSEUDO_SUB_PBA_FAIL_PROB = 0.10
PSEUDO_SAMPLING_SECONDS = 2.0


def pseudo_current(model_voltage: float) -> float:
    float(model_voltage)
    return random.uniform(0.0, 2.0)


def get_pf_status(model: str, avg_current: float, samples: Optional[List[float]] = None) -> str:
    if samples:
        try:
            post_startup = samples[5:] if len(samples) > 5 else []
            if post_startup and max(post_startup) > 4.0:
                return "FAIL(W748)"
        except ValueError:
            pass

        if len(samples) >= 3:
            critical_samples = samples[2:]
            if any(s < 0.01 for s in critical_samples):
                return "FAIL(W74A)" if random.random() < 0.85 else "FAIL(W748)"

    criteria = MODEL_CRITERIA.get(model, {})
    pass_thr = criteria.get("PASS", 1)
    f7p_thr = criteria.get("F7P")

    if f7p_thr is not None:
        if avg_current >= pass_thr:
            return "PASS"
        if avg_current >= f7p_thr:
            return "FAIL(W748)"
        return "FAIL(W74A)"

    if avg_current >= pass_thr:
        return "PASS"

    return "FAIL(W74A)" if random.random() < 0.85 else "FAIL(W748)"


def measure_current_and_get_avg_with_progress(
    model_voltage: float,
    progress_queue: queue.Queue,
    stop_event: threading.Event,
    prompt_response_queue: Optional[queue.Queue] = None,
    supply: Optional[Dict[str, str]] = None,
    series_store: Optional[Dict[str, Dict[str, List[float]]]] = None,
    series_lock: Optional[threading.Lock] = None,
    series_token: Optional[str] = None,
) -> None:
    supply_name = (supply or {}).get("name", "PSU")
    supply_addr = (supply or {}).get("address", DEFAULT_PSU_ADDRESS)

    if series_store is not None and series_lock is not None and series_token is not None:
        with series_lock:
            series_store[series_token] = {"times": [], "currents": []}

    def push_sample(sample_idx: int, current_val: float) -> None:
        if series_store is None or series_lock is None or series_token is None:
            return
        with series_lock:
            s = series_store.get(series_token)
            if not s:
                series_store[series_token] = {"times": [], "currents": []}
                s = series_store[series_token]
            s["times"].append(sample_idx)
            s["currents"].append(current_val)

    if USE_PSEUDO_CURRENT:
        _run_pseudo_measurement(
            model_voltage=model_voltage,
            supply_name=supply_name,
            progress_queue=progress_queue,
            stop_event=stop_event,
            prompt_response_queue=prompt_response_queue,
            push_sample=push_sample,
            series_token=series_token,
        )
        return

    _run_real_measurement(
        model_voltage=model_voltage,
        supply_name=supply_name,
        supply_addr=supply_addr,
        progress_queue=progress_queue,
        stop_event=stop_event,
        prompt_response_queue=prompt_response_queue,
        push_sample=push_sample,
        series_token=series_token,
    )


def _run_pseudo_measurement(
    model_voltage: float,
    supply_name: str,
    progress_queue: queue.Queue,
    stop_event: threading.Event,
    prompt_response_queue: Optional[queue.Queue],
    push_sample,
    series_token: Optional[str],
) -> None:
    try:
        progress_queue.put(("status", f"Using pseudo mode (supply {supply_name})"))
        progress_queue.put(("phase", "waiting"))

        try:
            prob = float(PSEUDO_SUB_PBA_FAIL_PROB)
        except Exception:
                prob = 0.0
        prob = max(0.0, min(1.0, prob))

        if random.random() < prob:
            if prompt_response_queue is None:
                progress_queue.put(("error", "Unable to confirm device connection. Please retry."))
                return

            progress_queue.put(("status", f"No current detected for {NO_CURRENT_SECONDS} seconds."))
            progress_queue.put(
                (
                    "prompt_device_check",
                    supply_name,
                    f"No current detected for {NO_CURRENT_SECONDS} seconds. Is a device connected to the power supply?",
                )
            )

            user_answer = _wait_for_prompt_answer(stop_event, prompt_response_queue)
            if user_answer:
                progress_queue.put(("sub_pba_fail", supply_name, series_token))
                return

        progress_queue.put(("status", "Simulating device connection..."))
        for _ in range(5):
            _raise_if_cancelled(stop_event)
            time.sleep(0.2)

        current_values: List[float] = []
        progress_queue.put(("phase", "measuring"))
        progress_queue.put(("status", "Simulating current measurements"))

        per_sample = float(PSEUDO_SAMPLING_SECONDS) / float(max(SAMPLE_COUNT, 1))
        sub_delay = per_sample / 5.0

        for i in range(SAMPLE_COUNT):
            _raise_if_cancelled(stop_event)
            current_value = pseudo_current(model_voltage)
            current_values.append(current_value)
            push_sample(i + 1, current_value)
            progress_queue.put(("tick", i + 1, SAMPLE_COUNT, current_value))

            for _ in range(5):
                _raise_if_cancelled(stop_event)
                time.sleep(sub_delay)

        last_5_avg = sum(current_values[-5:]) / 5 if len(current_values) >= 5 else sum(current_values) / len(current_values)
        total_avg = sum(current_values) / len(current_values)
        progress_queue.put(("done", last_5_avg, total_avg, supply_name, series_token))

    except Exception as e:
        progress_queue.put(("error", str(e)))


def _format_visa_connection_help(supply_name: str, supply_addr: str, err: Exception) -> str:
    return (
        f"Failed to connect to power supply.\n\n"
        f"Supply: {supply_name}\n"
        f"Address: {supply_addr}\n\n"
        f"Error: {err}\n\n"
        "Checklist:\n"
        "- Verify the address in AutoPowerTester_Configuration.json -> POWER_SUPPLIES\n"
        "- Verify NI-VISA or Keysight VISA is installed on this PC\n"
        "- Verify the instrument is powered on and connected (GPIB/USB/LAN)\n"
        "- Verify the GPIB interface and instrument address match\n"
    )


def _run_real_measurement(
    model_voltage: float,
    supply_name: str,
    supply_addr: str,
    progress_queue: queue.Queue,
    stop_event: threading.Event,
    prompt_response_queue: Optional[queue.Queue],
    push_sample,
    series_token: Optional[str],
) -> None:
    inst = None
    try:
        if not str(supply_addr).strip():
            raise ValueError(
                f"Power supply address is empty for {supply_name}. "
                "Please set POWER_SUPPLIES in AutoPowerTester_Configuration.json."
            )

        progress_queue.put(("status", f"Connecting to {supply_name} at {supply_addr}..."))

        rm = pyvisa.ResourceManager()
        inst = rm.open_resource(supply_addr)
        inst.write("OUTP ON")
        inst.write(f"VOLT {model_voltage}")

        progress_queue.put(("phase", "waiting"))
        progress_queue.put(("status", "Plug Anyway Jig into the phone."))

        zero_current_start: Optional[float] = None
        while True:
            _raise_if_cancelled(stop_event)

            current = float(inst.query("MEAS:CURR?").strip())
            if current > CURRENT_PRESENT_THRESHOLD_A:
                break

            if zero_current_start is None:
                zero_current_start = time.time()
            elif time.time() - zero_current_start >= NO_CURRENT_SECONDS:
                if prompt_response_queue is None:
                    progress_queue.put(("error", "Unable to confirm device connection. Please retry."))
                    return

                progress_queue.put(("status", f"No current detected for {NO_CURRENT_SECONDS} seconds."))
                progress_queue.put(
                    (
                        "prompt_device_check",
                        supply_name,
                        f"No current detected for {NO_CURRENT_SECONDS} seconds. Is a device connected to the power supply?",
                    )
                )

                user_answer = _wait_for_prompt_answer(stop_event, prompt_response_queue)
                if user_answer:
                    progress_queue.put(("sub_pba_fail", supply_name, series_token))
                    return

                zero_current_start = time.time()

            time.sleep(0.2)

        progress_queue.put(("phase", "measuring"))
        progress_queue.put(("status", "Measuring current"))

        current_values: List[float] = []
        for i in range(SAMPLE_COUNT):
            _raise_if_cancelled(stop_event)

            current_float = float(inst.query("MEAS:CURR?").strip())
            current_values.append(current_float)
            push_sample(i + 1, current_float)
            progress_queue.put(("tick", i + 1, SAMPLE_COUNT, current_float))
            time.sleep(1)

        last_5_avg = sum(current_values[-5:]) / 5 if len(current_values) >= 5 else sum(current_values) / len(current_values)
        total_avg = sum(current_values) / len(current_values)
        progress_queue.put(("done", last_5_avg, total_avg, supply_name, series_token))

    except Exception as e:
        progress_queue.put(("error", _format_visa_connection_help(supply_name, supply_addr, e)))

    finally:
        try:
            if inst is not None:
                inst.write("OUTP OFF")
                inst.close()
        except Exception:
            pass


def _raise_if_cancelled(stop_event: threading.Event) -> None:
    if stop_event.is_set():
        raise Exception("Measurement cancelled by user.")


def _wait_for_prompt_answer(stop_event: threading.Event, prompt_response_queue: queue.Queue) -> bool:
    while True:
        _raise_if_cancelled(stop_event)
        try:
            return bool(prompt_response_queue.get(timeout=0.1))
        except queue.Empty:
            continue


current_user = {"username": "", "is_admin": False, "is_dev": False}


def is_valid_imei(imei: str) -> bool:
    return imei.isdigit() and len(imei) == 15 and imei.startswith("3")


def find_panel_index_for_supply(name: str) -> Optional[int]:
    normalized = (name or "").strip()
    for i, ps in enumerate(POWER_SUPPLIES[:4]):
        if str(ps.get("name", "")).strip() == normalized:
            return i

    upper = normalized.upper()
    if upper.startswith("PSU"):
        digits = "".join(ch for ch in upper[3:] if ch.isdigit())
        if digits:
            try:
                return int(digits) - 1
            except Exception:
                return None

    return None


def prompt_no_current_detected(parent: tk.Misc, supply_name: str, question_en: str, anchor_widget=None) -> bool:
    supply_name = (supply_name or "PSU").strip() or "PSU"
    question_en = (question_en or "Is a device connected to the power supply?").strip() or "Is a device connected to the power supply?"
    question_es = "No se detectó corriente durante 15 segundos. ¿Hay un dispositivo conectado a la fuente de alimentación?"

    dialog_w, dialog_h = 656, 250

    win = tk.Toplevel(parent)
    win.title(f"No Current Detected - {supply_name}")
    win.geometry(f"{dialog_w}x{dialog_h}")
    win.resizable(False, False)
    try:
        win.transient(parent)
    except Exception:
        pass

    try:
        win.update_idletasks()
        if anchor_widget is not None:
            ax = int(anchor_widget.winfo_rootx())
            ay = int(anchor_widget.winfo_rooty())
            aw = int(anchor_widget.winfo_width())

            sx = int(win.winfo_screenwidth())
            sy = int(win.winfo_screenheight())
            margin = 10

            x = ax + aw + margin
            y = ay

            if x + dialog_w > sx - margin:
                x = ax - dialog_w - margin
            if x < margin:
                x = margin

            if y + dialog_h > sy - margin:
                y = sy - dialog_h - margin
            if y < margin:
                y = margin

            win.geometry(f"{dialog_w}x{dialog_h}+{x}+{y}")
    except Exception:
        pass

    container = tk.Frame(win, padx=14, pady=10)
    container.pack(fill="both", expand=True)

    tk.Label(container, text="NO CURRENT DETECTED", font=("TkDefaultFont", 16, "bold")).pack(anchor="w")
    tk.Label(container, text=f"PSU: {supply_name}", font=("TkDefaultFont", 14, "bold"), fg="red").pack(anchor="w", pady=(4, 8))

    tk.Label(container, text=question_en, font=("TkDefaultFont", 13), wraplength=620, justify="left").pack(anchor="w", pady=(0, 6))
    tk.Label(container, text=question_es, font=("TkDefaultFont", 13), wraplength=620, justify="left").pack(anchor="w", pady=(0, 10))

    btn_row = tk.Frame(container)
    btn_row.pack(anchor="center")

    result = {"value": False}

    def on_yes() -> None:
        result["value"] = True
        win.destroy()

    def on_no() -> None:
        result["value"] = False
        win.destroy()

    tk.Button(btn_row, text="YES / SÍ", font=("TkDefaultFont", 13, "bold"), width=10, command=on_yes).pack(side="left", padx=10)
    tk.Button(btn_row, text="NO", font=("TkDefaultFont", 13, "bold"), width=10, command=on_no).pack(side="left", padx=10)

    win.protocol("WM_DELETE_WINDOW", on_no)
    win.grab_set()
    try:
        win.focus_force()
    except Exception:
        pass

    parent.wait_window(win)
    return bool(result["value"])


def open_configuration_dialog(parent: tk.Misc) -> None:
    if not current_user["is_admin"]:
        messagebox.showinfo("Permission denied", "Only admin can open configuration.")
        return

    dialog = tk.Toplevel(parent)
    dialog.title("Configuration (Admin Only)")
    dialog.geometry(CONFIG_DIALOG_GEOMETRY)
    dialog.grab_set()

    nb = ttk.Notebook(dialog)
    nb.pack(fill="both", expand=True, padx=10, pady=10)

    tab_models = tk.Frame(nb)
    nb.add(tab_models, text="Models")

    tab_psu = tk.Frame(nb)
    nb.add(tab_psu, text="Power Supplies")

    tab_workers = tk.Frame(nb)
    nb.add(tab_workers, text="Workers")

    _build_models_tab(tab_models)
    _build_power_supplies_tab(tab_psu)
    _build_workers_tab(tab_workers)


def _build_models_tab(tab: tk.Frame) -> None:
    tab.grid_columnconfigure(2, weight=1)

    tk.Label(tab, text="Model:", font=("TkDefaultFont", 12)).grid(row=0, column=0, padx=10, pady=10, sticky="e")

    model_var = tk.StringVar()
    model_combo = ttk.Combobox(
        tab,
        textvariable=model_var,
        values=sorted(MODEL_VOLTAGE_MAP.keys()),
        state="readonly",
        font=("TkDefaultFont", 12),
        width=24,
    )
    model_combo.grid(row=0, column=1, padx=10, pady=10, sticky="w")

    tk.Label(tab, text="Or new model:", font=("TkDefaultFont", 12)).grid(row=1, column=0, padx=10, pady=5, sticky="e")
    new_model_entry = tk.Entry(tab, font=("TkDefaultFont", 12), width=26)
    new_model_entry.grid(row=1, column=1, padx=10, pady=5, sticky="w")

    tk.Label(tab, text="Boot On Voltage:", font=("TkDefaultFont", 12)).grid(row=2, column=0, padx=10, pady=5, sticky="e")
    voltage_entry = tk.Entry(tab, font=("TkDefaultFont", 12), width=26)
    voltage_entry.grid(row=2, column=1, padx=10, pady=5, sticky="w")

    tk.Label(tab, text="Criteria (one per line, e.g. PASS=0.8):", font=("TkDefaultFont", 12)).grid(
        row=3, column=0, padx=10, pady=5, sticky="ne"
    )
    criteria_text = tk.Text(tab, width=28, height=10, font=("TkDefaultFont", 12))
    criteria_text.grid(row=3, column=1, padx=10, pady=5, sticky="w")

    current_values_text = tk.Text(tab, width=60, height=18, font=("TkDefaultFont", 10), state="disabled")
    current_values_text.grid(row=0, column=2, rowspan=6, padx=(20, 10), pady=10, sticky="nsew")

    def format_criteria_for_display(criteria_dict: Dict[str, Any]) -> str:
        if not criteria_dict:
            return "No criteria"
        parts = []
        for key in sorted(criteria_dict.keys()):
            parts.append(f"{key}={criteria_dict[key]:g}")
        return ", ".join(parts)

    def refresh_models_display() -> None:
        lines = []
        for model_name in sorted(MODEL_VOLTAGE_MAP.keys()):
            voltage_value = MODEL_VOLTAGE_MAP.get(model_name, "")
            criteria_value = MODEL_CRITERIA.get(model_name, {})
            lines.append(f"{model_name} / {voltage_value} / {format_criteria_for_display(criteria_value)}")

        text_output = "\n".join(lines) if lines else "No models configured."
        current_values_text.config(state="normal")
        current_values_text.delete("1.0", tk.END)
        current_values_text.insert(tk.END, text_output)
        current_values_text.config(state="disabled")

    def load_model_settings(_event=None) -> None:
        if _event is not None:
            pass

        model = model_var.get()
        if not model:
            return

        new_model_entry.delete(0, tk.END)
        voltage_entry.delete(0, tk.END)
        voltage_entry.insert(0, str(MODEL_VOLTAGE_MAP.get(model, DEFAULT_BOOT_VOLTAGE)))

        criteria_text.delete("1.0", tk.END)
        criteria = MODEL_CRITERIA.get(model, {})
        for k, v in criteria.items():
            criteria_text.insert(tk.END, f"{k}={v}\n")

    def save_model_settings() -> None:
        model = new_model_entry.get().strip() or model_var.get().strip()
        if not model:
            messagebox.showwarning("Model Settings", "Please select or enter a model name.")
            return

        try:
            v = float(voltage_entry.get().strip())
        except ValueError:
            messagebox.showwarning("Model Settings", "Voltage must be a number.")
            return

        new_criteria: Dict[str, float] = {}
        for line in criteria_text.get("1.0", tk.END).splitlines():
            line = line.strip()
            if not line:
                continue
            if "=" not in line:
                messagebox.showwarning("Model Settings", f"Invalid criteria line: {line}")
                return

            key, val = line.split("=", 1)
            key = key.strip()
            try:
                val_float = float(val.strip())
            except ValueError:
                messagebox.showwarning("Model Settings", f"Invalid value in line: {line}")
                return
            new_criteria[key] = val_float

        MODEL_VOLTAGE_MAP[model] = v
        MODEL_CRITERIA[model] = new_criteria

        try:
            save_config(MODEL_VOLTAGE_MAP, MODEL_CRITERIA, WORKERS, POWER_SUPPLIES)
        except Exception as e:
            messagebox.showerror("Model Settings", f"Failed to save config:\n{e}")
            return

        model_combo["values"] = sorted(MODEL_VOLTAGE_MAP.keys())
        model_var.set(model)
        refresh_models_display()
        messagebox.showinfo("Model Settings", "Saved.")

    model_combo.bind("<<ComboboxSelected>>", load_model_settings)

    if MODEL_VOLTAGE_MAP:
        first_model = sorted(MODEL_VOLTAGE_MAP.keys())[0]
        model_var.set(first_model)
        load_model_settings()

    refresh_models_display()

    tk.Button(tab, text="Save Model", font=("TkDefaultFont", 12), command=save_model_settings).grid(
        row=4, column=0, columnspan=2, pady=10
    )


def _build_power_supplies_tab(tab: tk.Frame) -> None:
    tk.Label(tab, text="Supplies (name = address):", font=("TkDefaultFont", 12, "bold")).pack(padx=10, pady=(10, 5), anchor="w")

    supply_listbox = tk.Listbox(tab, font=("TkDefaultFont", 11), height=10)
    supply_listbox.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def refresh_supply_listbox() -> None:
        supply_listbox.delete(0, tk.END)
        for ps in POWER_SUPPLIES:
            supply_listbox.insert(tk.END, f"{ps.get('name', '')} = {ps.get('address', '')}")

    refresh_supply_listbox()

    entries = tk.Frame(tab)
    entries.pack(fill="x", padx=10, pady=5)

    tk.Label(entries, text="Name:", font=("TkDefaultFont", 11)).grid(row=0, column=0, padx=5, pady=5, sticky="e")
    ps_name_entry = tk.Entry(entries, font=("TkDefaultFont", 11))
    ps_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="we")

    tk.Label(entries, text="Address:", font=("TkDefaultFont", 11)).grid(row=1, column=0, padx=5, pady=5, sticky="e")
    ps_addr_entry = tk.Entry(entries, font=("TkDefaultFont", 11))
    ps_addr_entry.grid(row=1, column=1, padx=5, pady=5, sticky="we")

    entries.grid_columnconfigure(1, weight=1)

    def on_select_supply(_event=None) -> None:
        if _event is not None:
            pass

        sel = supply_listbox.curselection()
        if not sel:
            return

        idx = sel[0]
        text = supply_listbox.get(idx)
        if " = " in text:
            nm, ad = text.split(" = ", 1)
            ps_name_entry.delete(0, tk.END)
            ps_name_entry.insert(0, nm)
            ps_addr_entry.delete(0, tk.END)
            ps_addr_entry.insert(0, ad)

    supply_listbox.bind("<<ListboxSelect>>", on_select_supply)

    def add_or_update_supply() -> None:
        nm = ps_name_entry.get().strip()
        ad = ps_addr_entry.get().strip()
        if not nm:
            messagebox.showwarning("Power Supply Settings", "Name cannot be empty.")
            return

        if not ad:
            if not messagebox.askyesno(
                "Power Supply Address Empty",
                "Address is empty.\n\n"
                "The program will still run, but real measurements will not be able to connect.\n\n"
                "Do you want to save anyway?",
            ):
                return

        existing = next((ps for ps in POWER_SUPPLIES if ps.get("name") == nm), None)
        if existing:
            existing["address"] = ad
        else:
            if len(POWER_SUPPLIES) >= 4:
                messagebox.showwarning("Power Supply Settings", "Maximum 4 supplies allowed.")
                return
            POWER_SUPPLIES.append({"name": nm, "address": ad})

        try:
            save_config(MODEL_VOLTAGE_MAP, MODEL_CRITERIA, WORKERS, POWER_SUPPLIES)
        except Exception as e:
            messagebox.showerror("Power Supply Settings", f"Failed to save config:\n{e}")
            return

        refresh_supply_listbox()

    def delete_supply() -> None:
        sel = supply_listbox.curselection()
        if not sel:
            messagebox.showwarning("Power Supply Settings", "Please select a power supply to delete.")
            return

        idx = sel[0]
        text = supply_listbox.get(idx)
        nm = text.split(" = ", 1)[0] if " = " in text else text

        POWER_SUPPLIES[:] = [ps for ps in POWER_SUPPLIES if ps.get("name") != nm]

        try:
            save_config(MODEL_VOLTAGE_MAP, MODEL_CRITERIA, WORKERS, POWER_SUPPLIES)
        except Exception as e:
            messagebox.showerror("Power Supply Settings", f"Failed to save config:\n{e}")
            return

        refresh_supply_listbox()
        ps_name_entry.delete(0, tk.END)
        ps_addr_entry.delete(0, tk.END)

    btns = tk.Frame(tab)
    btns.pack(fill="x", padx=10, pady=10)

    tk.Button(btns, text="Add / Update", font=("TkDefaultFont", 11), command=add_or_update_supply).pack(side="left", padx=5)
    tk.Button(btns, text="Delete", font=("TkDefaultFont", 11), command=delete_supply).pack(side="left", padx=5)


def _build_workers_tab(tab: tk.Frame) -> None:
    tk.Label(tab, text="Workers (ID = password):", font=("TkDefaultFont", 12, "bold")).pack(padx=10, pady=(10, 5), anchor="w")

    worker_listbox = tk.Listbox(tab, font=("TkDefaultFont", 11), height=10)
    worker_listbox.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def refresh_worker_listbox() -> None:
        worker_listbox.delete(0, tk.END)
        for wid in sorted(WORKERS.keys()):
            worker_listbox.insert(tk.END, f"{wid} = {WORKERS[wid]}")

    refresh_worker_listbox()

    wf = tk.Frame(tab)
    wf.pack(fill="x", padx=10, pady=5)

    tk.Label(wf, text="Worker ID:", font=("TkDefaultFont", 11)).grid(row=0, column=0, padx=5, pady=5, sticky="e")
    worker_id_entry = tk.Entry(wf, font=("TkDefaultFont", 11))
    worker_id_entry.grid(row=0, column=1, padx=5, pady=5, sticky="we")

    tk.Label(wf, text="Password:", font=("TkDefaultFont", 11)).grid(row=1, column=0, padx=5, pady=5, sticky="e")
    worker_pw_entry = tk.Entry(wf, font=("TkDefaultFont", 11), show="*")
    worker_pw_entry.grid(row=1, column=1, padx=5, pady=5, sticky="we")

    wf.grid_columnconfigure(1, weight=1)

    def on_select_worker(_event=None) -> None:
        if _event is not None:
            pass

        sel = worker_listbox.curselection()
        if not sel:
            return

        idx = sel[0]
        text = worker_listbox.get(idx)
        if " = " in text:
            wid, pw = text.split(" = ", 1)
            worker_id_entry.delete(0, tk.END)
            worker_id_entry.insert(0, wid)
            worker_pw_entry.delete(0, tk.END)
            worker_pw_entry.insert(0, pw)

    worker_listbox.bind("<<ListboxSelect>>", on_select_worker)

    def add_or_update_worker() -> None:
        wid = worker_id_entry.get().strip()
        pw = worker_pw_entry.get().strip()

        if not wid or not pw:
            messagebox.showwarning("Worker Settings", "Worker ID and Password cannot be empty.")
            return

        if wid in (ADMIN_USERNAME, DEV_USERNAME):
            messagebox.showwarning("Worker Settings", "Cannot use reserved username as a worker ID.")
            return

        WORKERS[wid] = pw
        try:
            save_config(MODEL_VOLTAGE_MAP, MODEL_CRITERIA, WORKERS, POWER_SUPPLIES)
        except Exception as e:
            messagebox.showerror("Worker Settings", f"Failed to save config:\n{e}")
            return

        refresh_worker_listbox()

    def delete_worker() -> None:
        sel = worker_listbox.curselection()
        if not sel:
            messagebox.showwarning("Worker Settings", "Please select a worker to delete.")
            return

        idx = sel[0]
        text = worker_listbox.get(idx)
        if " = " not in text:
            return

        wid = text.split(" = ", 1)[0]

        if wid in WORKERS and messagebox.askyesno("Worker Settings", f"Delete worker '{wid}'?"):
            del WORKERS[wid]
            try:
                save_config(MODEL_VOLTAGE_MAP, MODEL_CRITERIA, WORKERS, POWER_SUPPLIES)
            except Exception as e:
                messagebox.showerror("Worker Settings", f"Failed to save config:\n{e}")
                return

            refresh_worker_listbox()
            worker_id_entry.delete(0, tk.END)
            worker_pw_entry.delete(0, tk.END)

    wb_frame = tk.Frame(tab)
    wb_frame.pack(fill="x", padx=10, pady=10)

    tk.Button(wb_frame, text="Add / Update", font=("TkDefaultFont", 11), command=add_or_update_worker).pack(side="left", padx=5)
    tk.Button(wb_frame, text="Delete", font=("TkDefaultFont", 11), command=delete_worker).pack(side="left", padx=5)


def center_window(win, width=None, height=None) -> None:
    try:
        win.update_idletasks()

        if width is None or height is None:
            w = int(win.winfo_width())
            h = int(win.winfo_height())
            if w <= 1 or h <= 1:
                w = int(win.winfo_reqwidth())
                h = int(win.winfo_reqheight())
        else:
            w = int(width)
            h = int(height)

        sw = int(win.winfo_screenwidth())
        sh = int(win.winfo_screenheight())
        x = max((sw - w) // 2, 0)
        y = max((sh - h) // 2, 0)
        win.geometry(f"{w}x{h}+{x}+{y}")
    except Exception:
        pass


@dataclass
class QueueMessageContext:
    tree: ttk.Treeview
    main_window: tk.Misc
    imei: str
    model: str
    progress_var: tk.StringVar
    progress_bar: ttk.Progressbar
    status_label: tk.Label
    finish_job_ui: Any
    panel_frames: List
    prompt_response_queue: queue.Queue


def _handle_status_msg(msg: Tuple, ctx: QueueMessageContext) -> Optional[str]:
    ctx.progress_var.set(msg[1])
    return None


def _handle_phase_msg(msg: Tuple, ctx: QueueMessageContext) -> Optional[str]:
    phase = msg[1]
    if phase == "waiting":
        ctx.progress_bar.config(mode="indeterminate", maximum=100, value=0)
        ctx.progress_bar.start(10)
        ctx.progress_var.set("Waiting for device current...")
    elif phase == "measuring":
        try:
            ctx.progress_bar.stop()
        except Exception:
            pass
        ctx.progress_bar.config(mode="determinate", maximum=SAMPLE_COUNT, value=0)
        if USE_PSEUDO_CURRENT:
            ctx.progress_var.set(f"Remaining: {PSEUDO_SAMPLING_SECONDS:.1f}s")
        else:
            ctx.progress_var.set(f"Remaining: {SAMPLE_COUNT}s")
    return None


def _handle_tick_msg(msg: Tuple, ctx: QueueMessageContext) -> Optional[str]:
    i, total, current_val = msg[1], msg[2], msg[3]
    ctx.progress_bar["value"] = i
    remaining = max(total - i, 0)
    if USE_PSEUDO_CURRENT:
        remaining_s = (float(PSEUDO_SAMPLING_SECONDS) * float(remaining)) / float(max(total, 1))
        ctx.progress_var.set(f"Remaining: {remaining_s:.1f}s  |  I={current_val:.3f}A")
    else:
        ctx.progress_var.set(f"Remaining: {remaining}s  |  I={current_val:.3f}A")
    return None


def _handle_prompt_device_check_msg(msg: Tuple, ctx: QueueMessageContext) -> Optional[str]:
    supply_for_prompt = msg[1] if len(msg) >= 3 else "PSU"
    question = msg[2] if len(msg) >= 3 else (msg[1] if len(msg) > 1 else "Is a device connected to the power supply?")

    idx = find_panel_index_for_supply(supply_for_prompt)
    anchor = ctx.panel_frames[idx] if idx is not None and 0 <= idx < len(ctx.panel_frames) else None

    answer = prompt_no_current_detected(ctx.main_window, supply_for_prompt, question, anchor_widget=anchor)
    ctx.prompt_response_queue.put(answer)
    return None


def _handle_done_msg(msg: Tuple, ctx: QueueMessageContext) -> Optional[str]:
    avg_last5 = msg[1]
    supply_used = msg[3]
    token = msg[4]

    with SERIES_LOCK:
        series = SERIES_BY_TOKEN.get(token, {"times": [], "currents": []})
        samples = list(series.get("times", []))
        currents = list(series.get("currents", []))

    try:
        avg_current = float(avg_last5)
        pf_status = get_pf_status(ctx.model, avg_current, samples=currents)
        id_val = current_user["username"]
        date_val = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        avg_current_str = f"{avg_current:.3f}A"

        row = (ctx.imei, ctx.model, avg_current_str, pf_status, id_val, supply_used, date_val)
        tag = "green_row" if pf_status == "PASS" else "red_row"
        item_id = ctx.tree.insert("", tk.END, values=row, tags=(tag,))
        RUN_SERIES_BY_ROW[item_id] = {"samples": samples, "currents": currents, "supply": supply_used}

        with LOG_WRITE_LOCK:
            try:
                detail_rows = [(ctx.imei, ctx.model, id_val, supply_used, date_val, s, c) for s, c in zip(samples, currents)]
                append_run_to_daily_log(summary_row=row, detail_rows=detail_rows)
            except PermissionError:
                show_log_file_open_error(ctx.main_window, get_daily_log_saved_path())
            except Exception as log_e:
                messagebox.showerror("Log Error", f"Failed to save log: {log_e}")

        ctx.progress_var.set(f"Done: {pf_status} ({avg_current:.3f}A)")
        ctx.status_label.config(text=pf_status, fg=("green" if pf_status == "PASS" else "red"))

    except Exception as e:
        messagebox.showerror("Measurement Error", f"Failed to process result: {e}")
        ctx.progress_var.set("Error")
    finally:
        ctx.finish_job_ui()
    return "stop"


def _handle_sub_pba_fail_msg(msg: Tuple, ctx: QueueMessageContext) -> Optional[str]:
    supply_used = msg[1]
    token = msg[2] if len(msg) > 2 else None

    with SERIES_LOCK:
        series = SERIES_BY_TOKEN.get(token, {"times": [], "currents": []}) if token else {"times": [], "currents": []}
        samples = list(series.get("times", []))
        currents = list(series.get("currents", []))

    try:
        pf_status = "FAIL(Sub PBA)"
        avg_current_str = "0.000A"
        id_val = current_user["username"]
        date_val = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row = (ctx.imei, ctx.model, avg_current_str, pf_status, id_val, supply_used, date_val)
        item_id = ctx.tree.insert("", tk.END, values=row, tags=("red_row",))
        RUN_SERIES_BY_ROW[item_id] = {
            "samples": samples or [0],
            "currents": currents or [0.0],
            "supply": supply_used,
        }

        with LOG_WRITE_LOCK:
            try:
                append_run_to_daily_log(
                    summary_row=row,
                    detail_rows=[(ctx.imei, ctx.model, id_val, supply_used, date_val, 0, 0.0)],
                )
            except PermissionError:
                show_log_file_open_error(ctx.main_window, get_daily_log_saved_path())
            except Exception as log_e:
                messagebox.showerror("Log Error", f"Failed to save log: {log_e}")

        ctx.progress_var.set("Done: FAIL(Sub PBA)")
        ctx.status_label.config(text="FAIL(Sub PBA)", fg="red")

    except Exception as e:
        messagebox.showerror("Measurement Error", f"Failed to log Sub PBA failure: {e}")
        ctx.progress_var.set("Error")
    finally:
        ctx.finish_job_ui()
    return "stop"


def _handle_error_msg(msg: Tuple, ctx: QueueMessageContext) -> Optional[str]:
    err = msg[1]
    messagebox.showerror("Measurement Error", f"{err}")
    ctx.progress_var.set("Error")
    ctx.status_label.config(text="ERROR", fg="red")
    ctx.finish_job_ui()
    return "stop"


MESSAGE_HANDLERS = {
    "status": _handle_status_msg,
    "phase": _handle_phase_msg,
    "tick": _handle_tick_msg,
    "prompt_device_check": _handle_prompt_device_check_msg,
    "done": _handle_done_msg,
    "sub_pba_fail": _handle_sub_pba_fail_msg,
    "error": _handle_error_msg,
}


def open_main_window(root: tk.Tk) -> None:
    main_window = tk.Toplevel(root)
    main_window.title(APP_TITLE_RESULTS)
    try:
        w_str, h_str = MAIN_WINDOW_GEOMETRY.split("x", 1)
        center_window(main_window, int(w_str), int(h_str))
    except Exception:
        main_window.geometry(MAIN_WINDOW_GEOMETRY)
        center_window(main_window)

    main_window.grid_rowconfigure(0, weight=1)
    main_window.grid_columnconfigure(0, weight=3, uniform="main")
    main_window.grid_columnconfigure(1, weight=2, uniform="main")

    columns = tuple(SUMMARY_HEADERS)

    style = ttk.Style(main_window)
    try:
        heading_font = tkfont.Font(name="ResultsHeadingFont", exists=False, font="TkHeadingFont")
    except Exception:
        heading_font = tkfont.Font(font="TkHeadingFont")
    heading_font.configure(size=int(RESULTS_HEADER_FONT_SIZE), weight="bold")
    style.configure("Results.Treeview.Heading", font=heading_font)

    try:
        body_font = tkfont.Font(name="ResultsBodyFont", exists=False, font="TkDefaultFont")
    except Exception:
        body_font = tkfont.Font(font="TkDefaultFont")
    body_font.configure(size=int(RESULTS_BODY_FONT_SIZE), weight="normal")
    style.configure("Results.Treeview", font=body_font)

    tree = ttk.Treeview(main_window, columns=columns, show="headings", selectmode="extended", style="Results.Treeview")
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=85, anchor="center")

    tree.column("Power Supply", width=70, anchor="center")
    tree.tag_configure("green_row", background="#b6fcd5")
    tree.tag_configure("red_row", background="#ffb6b6")

    scrollbar = ttk.Scrollbar(main_window, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)

    tree.grid(row=0, column=0, sticky="nsew")
    scrollbar.grid(row=0, column=0, sticky="nse")

    if not current_user.get("is_dev"):
        global USE_PSEUDO_CURRENT
        USE_PSEUDO_CURRENT = False

    def export_to_excel() -> None:
        rows = [tree.item(item)["values"] for item in tree.get_children()]
        if not rows:
            messagebox.showwarning("Export Data", "No data to export.")
            return

        df = pd.DataFrame(rows, columns=columns)
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        try:
            df.to_excel(file_path, index=False)
            messagebox.showinfo("Export Data", f"Data exported successfully to {file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export data.\n{e}")

    def delete_selected_rows() -> None:
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("Delete Row", "Please select one or more rows to delete.")
            return

        for item in selected:
            if item in RUN_SERIES_BY_ROW:
                del RUN_SERIES_BY_ROW[item]
            tree.delete(item)

    right = tk.Frame(main_window)
    right.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=10, pady=10)
    right.grid_columnconfigure(0, weight=1)
    right.grid_rowconfigure(2, weight=1)

    tk.Label(right, text="Run Panels (PSU1 - PSU4)", font=("TkDefaultFont", 14, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 8))

    panels_container = tk.Frame(right)
    panels_container.grid(row=1, column=0, sticky="nsew")
    panels_container.grid_columnconfigure(0, weight=1)
    panels_container.grid_columnconfigure(1, weight=1)

    bottom_area = tk.Frame(right)
    bottom_area.grid(row=2, column=0, sticky="nsew")
    bottom_area.grid_columnconfigure(0, weight=1)
    bottom_area.grid_rowconfigure(0, weight=1)
    bottom_area.grid_rowconfigure(1, weight=0)
    bottom_area.grid_rowconfigure(2, weight=1)

    center_bar = tk.Frame(bottom_area)
    center_bar.grid(row=1, column=0)
    center_bar.grid_columnconfigure(0, weight=1)

    export_btn = tk.Button(center_bar, text="Export to Excel", font=("TkDefaultFont", 12), command=export_to_excel)

    if current_user["is_admin"]:
        export_btn.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        tk.Button(center_bar, text="Delete Selected Row(s)", font=("TkDefaultFont", 12), command=delete_selected_rows).grid(
            row=1, column=0, sticky="ew", pady=(0, 8)
        )

        tk.Button(center_bar, text="Configuration", font=("TkDefaultFont", 12), command=lambda: open_configuration_dialog(main_window)).grid(
            row=2, column=0, sticky="ew"
        )
    else:
        export_btn.grid(row=0, column=0, sticky="ew")

    if current_user.get("is_dev"):
        pseudo_mode_var = tk.BooleanVar(value=USE_PSEUDO_CURRENT)

        def on_toggle_pseudo_mode() -> None:
            global USE_PSEUDO_CURRENT
            USE_PSEUDO_CURRENT = pseudo_mode_var.get()

        tk.Checkbutton(
            center_bar,
            text="Use pseudo current (no power supply)",
            font=("TkDefaultFont", 12),
            variable=pseudo_mode_var,
            command=on_toggle_pseudo_mode,
        ).grid(row=3, column=0, sticky="w", pady=(12, 0))

    panel_frames: List[Optional[ttk.LabelFrame]] = [None, None, None, None]

    def show_run_graph(item_id: str) -> None:
        data = RUN_SERIES_BY_ROW.get(item_id)
        if not data:
            messagebox.showwarning("Plot Current", "No measurement data stored for this run.")
            return

        samples = data.get("samples", [])
        currents = data.get("currents", [])
        if not samples or not currents:
            messagebox.showwarning("Plot Current", "Measurement data for this run is empty.")
            return

        currents_mA = [c * 1000.0 for c in currents]
        values = tree.item(item_id, "values") or ["", "", "", "", "", "", ""]

        imei = values[0]
        model = values[1]
        pf_status = values[3]
        supply_used = values[5]

        plot_win = tk.Toplevel(main_window)
        plot_win.title(f"Current vs Sample - IMEI {imei}")
        plot_win.geometry(PLOT_WINDOW_GEOMETRY)

        fig = Figure(figsize=(6, 4), dpi=100)
        ax = fig.add_subplot(111)
        ax.plot(samples, currents_mA, marker="o")
        ax.set_title(f"Current vs Sample\nIMEI: {imei}, Model: {model}, P/F: {pf_status}, PSU: {supply_used}")
        ax.set_xlabel("Sample #")
        ax.set_ylabel("Current (mA)")
        ax.grid(True)

        canvas = FigureCanvasTkAgg(fig, master=plot_win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    tree.bind("<Double-1>", lambda e: (lambda rid: show_run_graph(rid) if rid else None)(tree.identify_row(e.y)))

    def run_from_panel(
        panel_index: int,
        imei_entry: tk.Entry,
        model_entry: ttk.Combobox,
        progress_var: tk.StringVar,
        progress_bar: ttk.Progressbar,
        run_btn: tk.Button,
        cancel_btn: tk.Button,
        status_label: tk.Label,
    ) -> None:
        supply = POWER_SUPPLIES[panel_index] if panel_index < len(POWER_SUPPLIES) else {"name": f"PSU{panel_index + 1}", "address": ""}

        imei = imei_entry.get().strip()
        model = model_entry.get().strip()

        PANEL_STATE[panel_index]["last_model"] = model or PANEL_STATE[panel_index].get("last_model")
        PANEL_STATE[panel_index]["last_imei"] = imei or PANEL_STATE[panel_index].get("last_imei")

        if not is_valid_imei(imei) or not model:
            messagebox.showwarning("Input Error", "Please enter correct IMEI and Model.")
            return

        if model not in MODEL_VOLTAGE_MAP:
            messagebox.showwarning("Input Error", f"Unknown model: {model}")
            return

        if ACTIVE_JOBS[panel_index] is not None:
            messagebox.showwarning("Busy", f"{supply.get('name', 'PSU')} is already running. Please wait.")
            return

        if not USE_PSEUDO_CURRENT and not str(supply.get("address", "")).strip():
            messagebox.showwarning(
                "Power Supply Address Missing",
                f"{supply.get('name', 'PSU')} address is empty.\n\n"
                "Update AutoPowerTester_Configuration.json -> POWER_SUPPLIES.\n"
                "Or use DEV pseudo mode for testing without a power supply.",
            )
            return

        run_btn.config(state="disabled")
        imei_entry.config(state="disabled")
        model_entry.config(state="disabled")

        cancel_btn.config(state="normal")
        progress_bar.config(mode="indeterminate", maximum=100, value=0)
        progress_bar.start(10)
        progress_var.set("Initializing...")

        q: queue.Queue = queue.Queue()
        prompt_response_queue: queue.Queue = queue.Queue()
        stop_event = threading.Event()
        series_token = f"p{panel_index}-{int(time.time() * 1000)}"

        def on_cancel_inline() -> None:
            cancel_btn.config(state="disabled")
            stop_event.set()
            progress_var.set("Cancelling...")

        cancel_btn.config(command=on_cancel_inline)

        model_voltage = float(MODEL_VOLTAGE_MAP.get(model, DEFAULT_BOOT_VOLTAGE))
        worker = threading.Thread(
            target=measure_current_and_get_avg_with_progress,
            args=(
                model_voltage,
                q,
                stop_event,
                prompt_response_queue,
                supply,
                SERIES_BY_TOKEN,
                SERIES_LOCK,
                series_token,
            ),
            daemon=True,
        )
        worker.start()

        ACTIVE_JOBS[panel_index] = PanelJob(panel_index=panel_index, thread=worker, stop_event=stop_event, series_token=series_token)

        def finish_job_ui() -> None:
            ACTIVE_JOBS[panel_index] = None
            try:
                progress_bar.stop()
            except Exception:
                pass

            progress_bar.config(mode="determinate", maximum=SAMPLE_COUNT, value=0)
            cancel_btn.config(state="disabled")
            run_btn.config(state="normal")
            imei_entry.config(state="normal")
            model_entry.config(state="normal")

        def poll_queue_inline() -> None:
            ctx = QueueMessageContext(
                tree=tree,
                main_window=main_window,
                imei=imei,
                model=model,
                progress_var=progress_var,
                progress_bar=progress_bar,
                status_label=status_label,
                finish_job_ui=finish_job_ui,
                panel_frames=panel_frames,
                prompt_response_queue=prompt_response_queue,
            )

            try:
                while True:
                    msg = q.get_nowait()
                    kind = msg[0]

                    handler = MESSAGE_HANDLERS.get(kind)
                    if handler:
                        result = handler(msg, ctx)
                        if result == "stop":
                            return

            except queue.Empty:
                pass

            main_window.after(100, poll_queue_inline)

        poll_queue_inline()

    def create_panel(parent: tk.Misc, panel_index: int) -> ttk.LabelFrame:
        supply = POWER_SUPPLIES[panel_index] if panel_index < len(POWER_SUPPLIES) else {"name": f"PSU{panel_index + 1}", "address": ""}

        frame = ttk.LabelFrame(parent, text=f"{supply.get('name', f'PSU{panel_index + 1}')}", padding=8)
        frame.grid_columnconfigure(1, weight=1)

        tk.Label(frame, text="IMEI:", font=("TkDefaultFont", 12)).grid(row=0, column=0, sticky="e", padx=6, pady=6)
        imei_entry = tk.Entry(frame, font=("TkDefaultFont", 12))
        imei_entry.grid(row=0, column=1, sticky="ew", padx=6, pady=6)

        tk.Label(frame, text="Model:", font=("TkDefaultFont", 12)).grid(row=1, column=0, sticky="e", padx=6, pady=6)
        model_var = tk.StringVar()
        model_entry = ttk.Combobox(frame, textvariable=model_var, values=sorted(MODEL_VOLTAGE_MAP.keys()), state="normal", font=("TkDefaultFont", 12))
        model_entry.grid(row=1, column=1, sticky="ew", padx=6, pady=6)

        def apply_model_filter(_event=None) -> None:
            if _event is not None:
                pass

            typed = model_var.get().strip().lower()
            all_models = sorted(MODEL_VOLTAGE_MAP.keys())
            if not typed:
                model_entry["values"] = all_models
                return

            filtered = [m for m in all_models if typed in m.lower()]
            model_entry["values"] = filtered if filtered else all_models
            if filtered:
                try:
                    model_entry.event_generate("<Down>")
                except Exception:
                    pass

        model_entry.bind("<KeyRelease>", apply_model_filter)

        last_model = PANEL_STATE[panel_index].get("last_model")
        if last_model:
            model_var.set(last_model)

        last_imei = PANEL_STATE[panel_index].get("last_imei")
        if last_imei:
            imei_entry.insert(0, last_imei)

        progress_var = tk.StringVar(value="Idle")
        tk.Label(frame, textvariable=progress_var, font=("TkDefaultFont", 10), anchor="w").grid(
            row=2, column=0, columnspan=2, sticky="ew", padx=6, pady=(2, 2)
        )

        progress_bar = ttk.Progressbar(frame, mode="determinate", maximum=SAMPLE_COUNT, value=0, length=240)
        progress_bar.grid(row=3, column=0, columnspan=2, sticky="ew", padx=6, pady=(0, 6))

        status_label = tk.Label(frame, text="", font=("TkDefaultFont", 11, "bold"), anchor="w")
        status_label.grid(row=4, column=0, columnspan=2, sticky="ew", padx=6, pady=(0, 6))

        btn_row = tk.Frame(frame)
        btn_row.grid(row=5, column=0, columnspan=2, pady=(4, 2))

        cancel_btn = tk.Button(btn_row, text="Cancel", font=("TkDefaultFont", 11), state="disabled")
        cancel_btn.pack(side="left", padx=4)

        def do_run() -> None:
            status_label.config(text="", fg="black")
            run_from_panel(panel_index, imei_entry, model_entry, progress_var, progress_bar, run_btn, cancel_btn, status_label)

        def on_return(event=None) -> None:
            del event
            do_run()

        imei_entry.bind("<Return>", on_return)
        model_entry.bind("<Return>", on_return)

        run_btn = tk.Button(btn_row, text="RUN", font=("TkDefaultFont", 14, "bold"), width=10, command=do_run)
        run_btn.pack(side="left", padx=4)

        return frame

    p1 = create_panel(panels_container, 0)
    p2 = create_panel(panels_container, 1)
    p3 = create_panel(panels_container, 2)
    p4 = create_panel(panels_container, 3)

    panel_frames[0] = p1
    panel_frames[1] = p2
    panel_frames[2] = p3
    panel_frames[3] = p4

    p1.grid(row=0, column=0, sticky="ew", padx=6, pady=6)
    p2.grid(row=0, column=1, sticky="ew", padx=6, pady=6)
    p3.grid(row=1, column=0, sticky="ew", padx=6, pady=6)
    p4.grid(row=1, column=1, sticky="ew", padx=6, pady=6)

    def on_close() -> None:
        for job in list(ACTIVE_JOBS):
            if job is None:
                continue
            try:
                job.stop_event.set()
            except Exception:
                pass

        for job in list(ACTIVE_JOBS):
            if job is None:
                continue
            try:
                if job.thread.is_alive():
                    job.thread.join(timeout=2)
            except Exception:
                pass

        for i in range(len(ACTIVE_JOBS)):
            ACTIVE_JOBS[i] = None

        main_window.destroy()
        root.destroy()

    main_window.protocol("WM_DELETE_WINDOW", on_close)


def try_login(root: tk.Tk, entry_username: tk.Entry, entry_password: tk.Entry) -> None:
    username = entry_username.get().strip()
    password = entry_password.get().strip()

    global USE_PSEUDO_CURRENT

    if username == DEV_USERNAME and password == DEV_PASSWORD:
        current_user["username"] = username
        current_user["is_admin"] = True
        current_user["is_dev"] = True
        messagebox.showinfo("Login", "Login successful!")
        root.withdraw()
        open_main_window(root)
        return

    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        current_user["username"] = username
        current_user["is_admin"] = True
        current_user["is_dev"] = False
        USE_PSEUDO_CURRENT = False
        messagebox.showinfo("Login", "Login successful!")
        root.withdraw()
        open_main_window(root)
        return

    if username in WORKERS and password == WORKERS.get(username):
        current_user["username"] = username
        current_user["is_admin"] = False
        current_user["is_dev"] = False
        USE_PSEUDO_CURRENT = False
        messagebox.showinfo("Login", "Login successful!")
        root.withdraw()
        open_main_window(root)
        return

    messagebox.showerror("Login", "Invalid credentials.")


def main() -> None:
    root = tk.Tk()
    root.title(APP_TITLE_LOGIN)

    tk.Label(root, text="Username:", font=("TkDefaultFont", 26)).grid(row=0, column=0, padx=50, pady=50)
    entry_username = tk.Entry(root, font=("TkDefaultFont", 26))
    entry_username.grid(row=0, column=1, padx=30, pady=30)

    tk.Label(root, text="Password:", font=("TkDefaultFont", 26)).grid(row=1, column=0, padx=50, pady=50)
    entry_password = tk.Entry(root, show="*", font=("TkDefaultFont", 26))
    entry_password.grid(row=1, column=1, padx=30, pady=30)

    tk.Button(root, text="Login", font=("TkDefaultFont", 20), command=lambda: try_login(root, entry_username, entry_password)).grid(
        row=2, column=0, columnspan=2, padx=10, pady=10
    )

    center_window(root)

    root.mainloop()


if __name__ == "__main__":
    main()
